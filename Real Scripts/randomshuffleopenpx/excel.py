import random
from openpyxl import load_workbook



def do_caculate(tmpdata):
    result_set = [0 for _ in range(row//GROUP)]
    for i in range(len(result_set)):
        temp_max = max(max(tmpdata[i*3][1:column+1]),max(tmpdata[i*3+1][1:column+1]),max(tmpdata[i*3+2][1:column+1]))
        temp_min = min(min(tmpdata[i*3][1:column+1]),min(tmpdata[i*3+1][1:column+1]),min(tmpdata[i*3+2][1:column+1]))
        result_set[i] = round(temp_max - temp_min,3)
        if result_set[i] > 0.018:
            return 9999999
    return round(sum(result_set),3)


ITERATE_TIME = 1000000
GROUP = 3
FILE_PATH = 'CSB-G.xlsx'

min_res = 99999999999999999999

wb = load_workbook(FILE_PATH)
sheet = wb.worksheets[0]
row = sheet.max_row - 1
column = sheet.max_column - 1

data = [[0 for _ in range(column + 1)] for _ in range(row)]


for i in range(row):
    for j in range(column+1):
        data[i][j] = sheet.cell(row=i+2, column=j+1).value

output_data = data[:]


for _ in range(ITERATE_TIME):
    tmp_num =  do_caculate(data)
    if min_res > tmp_num:
        min_res = tmp_num
        output_data = data[:]

    random.shuffle(data)
    

output_ws = wb.create_sheet("Mysheet",1)
output_ws.cell(row=1, column=1).value = '编号'

for i in range(row):
    for j in range(column+1):
        if i == 0 and j > 0:
            output_ws.cell(row=i+1, column=j+1).value = j
        
        output_ws.cell(row=i+2, column=j+1).value = output_data[i][j]

wb.save(FILE_PATH)
print(min_res)
